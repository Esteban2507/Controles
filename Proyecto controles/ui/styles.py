"""Módulo de estilos visuales."""
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# =========================
# Colores
# =========================
COLOR_HEADER_BG = "2C3E50"
COLOR_HEADER_FONT = "FFFFFF"
COLOR_ALTERNATE_ROW = "F2F4F4"

COLOR_OK_FILL = "D5F5E3"
COLOR_OK_FONT = "196F3D"
COLOR_ERROR_FILL = "FADBD8"
COLOR_ERROR_FONT = "943126"
COLOR_WARN_FILL = "FCF3CF"
COLOR_WARN_FONT = "9A7D0A"
COLOR_INFO_FILL = "D6EAF8"
COLOR_INFO_FONT = "21618C"

BORDER_THIN = Side(style="thin", color="BDC3C7")
BORDER_MEDIUM = Side(style="medium", color="7F8C8D")

ESTILO_BORDE_TABLA = Border(left=BORDER_THIN, right=BORDER_THIN, top=BORDER_THIN, bottom=BORDER_THIN)
ESTILO_BORDE_HEADER = Border(left=BORDER_THIN, right=BORDER_THIN, top=BORDER_MEDIUM, bottom=BORDER_MEDIUM)


def apply_header_style(ws, min_row, max_row, min_col, max_col):
    """Aplica estilo de encabezado a un rango de celdas."""
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = ESTILO_BORDE_TABLA
            if r == min_row:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")


def get_status_color(status, colors_cfg):
    """Obtiene el color hexadecimal para un status."""
    default_colors = {
        "OK": "00B050",
        "Error de Monto": "FF99CC",
        "Error de Moneda": "FFFF00",
        "Error de Tipo Cambio": "FFB347",
        "Duplicado en ARCA": "D6EAF8",
        "Faltante en ARCA": "FCF3CF",
        "Posible duplicado": "FFC000",
    }
    
    color_hex = colors_cfg.get(status, default_colors.get(status, "CCCCCC"))
    return "#" + color_hex.replace("#", "")
