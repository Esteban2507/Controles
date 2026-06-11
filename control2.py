import logging
import re
import unicodedata
from pathlib import Path
import time
from datetime import timedelta
import sys
import threading
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure

# =========================
# Colores y Estilos
# =========================
COLOR_HEADER_BG   = "2C3E50"
COLOR_HEADER_FONT = "FFFFFF"
COLOR_ALTERNATE_ROW = "F2F4F4"

COLOR_OK_FILL    = "D5F5E3"
COLOR_OK_FONT    = "196F3D"
COLOR_ERROR_FILL = "FADBD8"
COLOR_ERROR_FONT = "943126"
COLOR_WARN_FILL  = "FCF3CF"
COLOR_WARN_FONT  = "9A7D0A"
COLOR_INFO_FILL  = "D6EAF8"
COLOR_INFO_FONT  = "21618C"

BORDER_THIN   = Side(style="thin",   color="BDC3C7")
BORDER_MEDIUM = Side(style="medium", color="7F8C8D")

ESTILO_BORDE_TABLA  = Border(left=BORDER_THIN, right=BORDER_THIN, top=BORDER_THIN, bottom=BORDER_THIN)
ESTILO_BORDE_HEADER = Border(left=BORDER_THIN, right=BORDER_THIN, top=BORDER_MEDIUM, bottom=BORDER_MEDIUM)

LOG_FILE = "control_facturas.log"


# =========================
# Infraestructura
# =========================
def setup_logging(level_name="WARNING"):
    level = getattr(logging, level_name.upper(), logging.WARNING)
    logging.basicConfig(
        level=level,
        format="%(asctime)s - %(levelname)s - %(message)s",
        handlers=[
            logging.FileHandler(LOG_FILE, encoding="utf-8"),
            logging.StreamHandler(),
        ],
    )

def cargar_config():
    if hasattr(sys, '_MEIPASS'):
        base_path = Path(sys.executable).parent
    else:
        base_path = Path(__file__).resolve().parent
        
    cfg_path = base_path / "config.yaml"
    if not cfg_path.exists():
        cfg_path = Path.cwd() / "config.yaml"
    if not cfg_path.exists():
        raise FileNotFoundError(f"No se encontró 'config.yaml' en {base_path} ni en {Path.cwd()}")
    with open(cfg_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)

class ProgressRedirector(object):
    def __init__(self, string_var, progress_bar, root):
        self.string_var = string_var
        self.progress_bar = progress_bar
        self.root = root

    def write(self, string):
        s = string.strip()
        if s:
            self.root.after(0, self._update_gui, s)

    def _update_gui(self, s):
        if len(s) > 100:
            s_short = s[:97] + "..."
        else:
            s_short = s
            
        match = re.search(r"(\d)/8", s_short)
        if match:
            step = int(match.group(1))
            self.progress_bar["value"] = (step / 8.0) * 100
            self.string_var.set(s_short)
        else:
            if ">>>" not in s_short:
                self.string_var.set(s_short)

    def flush(self):
        pass


# Utilidades de datos
# =========================
def convertir_monto(serie: pd.Series) -> pd.Series:
    if pd.api.types.is_numeric_dtype(serie):
        return serie.fillna(0)
    s = serie.astype(str).str.replace(".", "", regex=False).str.replace(",", ".", regex=False)
    return pd.to_numeric(s, errors="coerce").fillna(0)

def normalizar_moneda(valor, equivalencias):
    if pd.isna(valor) or valor is None:
        return ""
    v = str(valor).strip().upper()
    if v in equivalencias.get("ARCA", {}):
        return equivalencias["ARCA"][v]
    if v in equivalencias.get("genericas", {}):
        return equivalencias["genericas"][v]
    return v

def normalizar_duplicados(serie: pd.Series) -> pd.Series:
    return serie.astype(str).str.strip()

def parsear_fecha_columna(df: pd.DataFrame, col: str) -> pd.DataFrame:
    if col not in df.columns:
        return df
    df[col] = pd.to_datetime(df[col], errors="coerce").dt.date
    return df

def canonicalize(label: str) -> str:
    if label is None:
        return ""
    s = str(label)
    s = unicodedata.normalize("NFD", s).encode("ascii", "ignore").decode("ascii")
    s = s.lower()
    s = re.sub(r'[^a-z0-9]+', '', s)
    return s

def detectar_candidatas_amount(df: pd.DataFrame):
    cands = []
    for col in df.columns:
        c = canonicalize(col)
        if c in ("monto", "montototal", "imptotal", "imptotal2", "amount"):
            cands.append(col)
    seen, res = set(), []
    for c in cands:
        if c not in seen:
            res.append(c); seen.add(c)
    return res

def seleccionar_mejor_amount(df: pd.DataFrame) -> str | None:
    cands = detectar_candidatas_amount(df)
    if not cands:
        return None
    mejor, mejor_score = None, -1
    for col in cands:
        vals = convertir_monto(df[col])
        score = vals.notna().sum()
        if score > mejor_score:
            mejor, mejor_score = col, score
    if len(cands) > 1 and mejor is not None:
        logging.info(f"Varias columnas de monto {cands}. Usando '{mejor}'.")
    return mejor

def seleccionar_mejor_currency(df: pd.DataFrame) -> str | None:
    candidatos = [c for c in df.columns if canonicalize(c) in ("moneda", "moneda2", "currency")]
    if not candidatos:
        return None
    for pref in ["Moneda", "currency", "Moneda 2"]:
        if pref in df.columns:
            s = df[pref].astype(str).str.strip()
            if s.notna().any() and (s != "").any():
                return pref
    return candidatos[0]

def rename_robusto(df: pd.DataFrame, mapping: dict, sistema: str, obligatorias: list) -> pd.DataFrame:
    canon_map = {canonicalize(k): v for k, v in mapping.items()}
    out_rename = {}
    for col in df.columns:
        c = canonicalize(col)
        if c in canon_map:
            destino = canon_map[c]
            if destino in out_rename.values():
                continue
            out_rename[col] = destino
    df = df.rename(columns=out_rename)

    # Heurística duplicados
    if "duplicados" in obligatorias and "duplicados" not in df.columns:
        candidatos = []
        for col in df.columns:
            c = canonicalize(col)
            if any(p in c for p in ["duplic", "ctrldupli", "concatenadomonto", "concatenadomasmonto", "concatenadomont"]):
                candidatos.append(col)
        if len(candidatos) == 1:
            df = df.rename(columns={candidatos[0]: "duplicados"})
            logging.info(f"[{sistema}] 'duplicados' por heurística: {candidatos[0]}")
        elif len(candidatos) > 1:
            elegido = max(candidatos, key=len)
            df = df.rename(columns={elegido: "duplicados"})
            logging.warning(f"[{sistema}] 'duplicados' múltiples {candidatos}. Usado: {elegido}")

    # Heurística vendor
    if "vendor" in obligatorias and "vendor" not in df.columns:
        candidatos = []
        for col in df.columns:
            c = canonicalize(col)
            if any(p in c for p in ["receptor", "proveedor", "vendor", "denominacion"]):
                candidatos.append(col)
        if len(candidatos) == 1:
            df = df.rename(columns={candidatos[0]: "vendor"})
            logging.info(f"[{sistema}] 'vendor' por heurística: {candidatos[0]}")
        elif len(candidatos) > 1:
            elegido = max(candidatos, key=len)
            df = df.rename(columns={elegido: "vendor"})
            logging.warning(f"[{sistema}] 'vendor' múltiples {candidatos}. Usado: {elegido}")

    faltantes = [req for req in obligatorias if req not in df.columns]
    if faltantes:
        raise KeyError(
            f"En {sistema} faltan columnas requeridas: {faltantes}. "
            f"Encabezados disponibles: {list(df.columns)}. "
            f"Revisá 'config.yaml' o los encabezados del archivo."
        )
    return df

def assert_cols(df: pd.DataFrame, sistema: str, requeridas: list):
    faltan = [c for c in requeridas if c not in df.columns]
    if faltan:
        raise KeyError(
            f"En {sistema} faltan columnas requeridas: {faltan}. "
            f"Cabeceras cargadas: {list(df.columns)}. Revisá config.yaml o los encabezados."
        )


# =========================
# Estilos y Dashboard
# =========================
def aplicar_estilo_tabla(ws, title=None, freeze_cell="G2"):
    max_col = ws.max_column
    max_row = ws.max_row
    # Encabezados
    fill_header = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
    font_header = Font(color=COLOR_HEADER_FONT, bold=True, size=11)
    for cell in ws[1]:
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = ESTILO_BORDE_HEADER
    ws.row_dimensions[1].height = 25
    # Cuerpo
    for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        is_even = (row[0].row % 2 == 0)
        fill_row = PatternFill(start_color=COLOR_ALTERNATE_ROW, end_color=COLOR_ALTERNATE_ROW, fill_type="solid") if is_even else None
        for cell in row:
            cell.border = ESTILO_BORDE_TABLA
            if fill_row:
                cell.fill = fill_row
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right")
                # Detectar si la columna se llama amount, Diferencia o monto para dar formato moneda
                col_title = str(ws.cell(row=1, column=cell.column).value).lower()
                if any(x in col_title for x in ["amount", "monto", "diferencia"]):
                    cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            else:
                cell.alignment = Alignment(horizontal="left")
    if not ws._tables and ws.max_row > 1 and ws.max_column > 0:
        table_name = f"Tabla_{ws.title.replace(' ', '_')}"
        table = Table(displayName=table_name, ref=ws.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
    # Filtros, freeze y ancho
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = freeze_cell
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[letter].width = min((max_len + 2) * 1.1, 60)

def aplicar_formato_condicional_avanzado(ws):
    status_col_let = None
    for cell in ws[1]:
        if str(cell.value).lower() == "status":
            status_col_let = get_column_letter(cell.column)
            break
    if not status_col_let or ws.max_row < 2:
        return
    rng = f"{status_col_let}2:{status_col_let}{ws.max_row}"
    ws.conditional_formatting.add(
        rng,
        FormulaRule(
            formula=[f'${status_col_let}2="OK"'],
            font=Font(color=COLOR_OK_FONT, bold=True),
            fill=PatternFill(start_color=COLOR_OK_FILL, end_color=COLOR_OK_FILL, fill_type="solid")
        )
    )
    ws.conditional_formatting.add(
        rng,
        FormulaRule(
            formula=[f'ISNUMBER(SEARCH("Error",{status_col_let}2))'],
            font=Font(color=COLOR_ERROR_FONT, bold=True),
            fill=PatternFill(start_color=COLOR_ERROR_FILL, end_color=COLOR_ERROR_FILL, fill_type="solid")
        )
    )
    ws.conditional_formatting.add(
        rng,
        FormulaRule(
            formula=[f'ISNUMBER(SEARCH("Faltante",{status_col_let}2))'],
            font=Font(color=COLOR_WARN_FONT, bold=True),
            fill=PatternFill(start_color=COLOR_WARN_FILL, end_color=COLOR_WARN_FILL, fill_type="solid")
        )
    )
    ws.conditional_formatting.add(
        rng,
        FormulaRule(
            formula=[f'ISNUMBER(SEARCH("Duplicado",{status_col_let}2))'],
            font=Font(color=COLOR_INFO_FONT, bold=True),
            fill=PatternFill(start_color=COLOR_INFO_FILL, end_color=COLOR_INFO_FILL, fill_type="solid")
        )
    )

def AplicarEstiloTablaResumen(ws, min_r, max_r, min_c, max_c):
    for r in range(min_r, max_r + 1):
        for c in range(min_c, max_c + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = ESTILO_BORDE_TABLA
            if r == min_r:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")

def generar_dashboard(wb, df_resumen, df_detalle, config, exec_msg=None):
    if "Dashboard" in wb.sheetnames:
        ws = wb["Dashboard"]
    else:
        ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False

    ws["B2"] = "DASHBOARD DE CONTROL DE FACTURAS"
    ws["B2"].font = Font(size=20, bold=True, color=COLOR_HEADER_BG)
    ws.merge_cells("B2:J2")

    total_regs = len(df_detalle)
    total_ok   = len(df_detalle[df_detalle["Status"] == "OK"])
    total_err  = total_regs - total_ok
    pct_ok     = (total_ok / total_regs) if total_regs else 0

    kpi_start_row = 4
    kpis = [
        ("Total Registros", total_regs, "E8DAEF"),
        ("Correctos (OK)", total_ok, COLOR_OK_FILL),
        ("Con Errores/Alertas", total_err, COLOR_ERROR_FILL),
        ("% Efectividad", f"{pct_ok:.1%}", "D6EAF8"),
    ]
    col = 2
    for title, val, color in kpis:
        cell_title = ws.cell(row=kpi_start_row, column=col)
        cell_val   = ws.cell(row=kpi_start_row+1, column=col)
        cell_title.value = title
        cell_title.font  = Font(bold=True, size=12, color="555555")
        cell_title.alignment = Alignment(horizontal="center")
        cell_val.value = val
        cell_val.font = Font(bold=True, size=16)
        cell_val.alignment = Alignment(horizontal="center")
        cell_val.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for r in range(kpi_start_row, kpi_start_row+2):
            ws.cell(row=r, column=col).border = ESTILO_BORDE_TABLA
        col += 2

    if exec_msg:
        ws["B6"] = "Tiempo de ejecución"
        ws["C6"] = exec_msg
        ws["B6"].font = Font(bold=True, color="555555")
        ws["C6"].font = Font(bold=False, color="555555")

    ws["B8"] = "Resumen por Estado"
    ws["B8"].font = Font(bold=True, size=14, color=COLOR_HEADER_BG)
    start_row_data = 9
    ws.cell(row=start_row_data, column=2, value="Estado")
    ws.cell(row=start_row_data, column=3, value="Cantidad")
    for idx, row in df_resumen.iterrows():
        r = start_row_data + 1 + idx
        ws.cell(row=r, column=2, value=row["Status"])
        ws.cell(row=r, column=3, value=row["Cantidad"])
    AplicarEstiloTablaResumen(ws, start_row_data, start_row_data + len(df_resumen), 2, 3)

    pie = PieChart()
    pie.title = "Distribución de Estados"
    pie.width = 15; pie.height = 10
    labels = Reference(ws, min_col=2, min_row=start_row_data+1, max_row=start_row_data+len(df_resumen))
    data   = Reference(ws, min_col=3, min_row=start_row_data,   max_row=start_row_data+len(df_resumen))
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    
    colors_cfg = config.get("dashboard_colors", {})
    default_colors = {"OK": "00B050", "Error de Monto": "FF99CC", "Error de Moneda": "FFFF00", "Duplicado en ARCA": "D6EAF8"}
    
    for i, status in enumerate(df_resumen["Status"]):
        color_hex = colors_cfg.get(status, default_colors.get(status, "CCCCCC")).replace("#", "")
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color_hex
        pie.series[0].dPt.append(pt)
        
    ws.add_chart(pie, "E8")

    bar = BarChart()
    bar.type = "col"; bar.style = 10
    bar.title = "Cantidad por Estado"
    bar.y_axis.title = 'Registros'
    bar.x_axis.title = 'Estado'
    bar.width = 15; bar.height = 10
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(labels)
    
    for i, status in enumerate(df_resumen["Status"]):
        color_hex = colors_cfg.get(status, default_colors.get(status, "CCCCCC")).replace("#", "")
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color_hex
        bar.series[0].dPt.append(pt)
        
    ws.add_chart(bar, "L8")


# =========================
# Lectura robusta
# =========================
def read_excel_robust(file_path, sheet_name_pref):
    try:
        return pd.read_excel(file_path, sheet_name=sheet_name_pref, engine="openpyxl")
    except ValueError:
        msg = f"Hoja '{sheet_name_pref}' no encontrada en '{Path(file_path).name}'. Leyendo la primera hoja."
        print(f" [AVISO] {msg}")
        logging.warning(msg)
        return pd.read_excel(file_path, sheet_name=0, engine="openpyxl")


# =========================
# Lógica principal
# =========================
def process_data(config, arca_path, cdp_path, e1_path="", fecha_filtro=None):
    diagnostico = bool(config.get("diagnostico", False))
    print(f">>> Diagnóstico: {diagnostico}")

    # === CRONÓMETRO ===
    t0 = time.perf_counter()

    # 1) Lectura
    print("1/8 Leyendo ARCA (hoja 'AFIP' o primera)...")
    df_arca_raw = read_excel_robust(arca_path, "AFIP")
    print(f"   ARCA filas={len(df_arca_raw):,} cols={len(df_arca_raw.columns)}")

    print("2/8 Leyendo CDP (hoja 'CDP' o primera)...")
    if cdp_path:
        df_cdp = read_excel_robust(cdp_path, "CDP")
        print(f"   CDP  filas={len(df_cdp):,} cols={len(df_cdp.columns)}")
        
        # Detectar columna de posting (se usará como key de historial)
        posting_col = None
        for col in df_cdp.columns:
            lc = str(col).lower()
            if "posting" in lc or ("fecha" in lc and "posting" in lc):
                posting_col = col
                break
        if posting_col is not None:
            try:
                df_cdp["posting_date_CDP"] = pd.to_datetime(df_cdp[posting_col], errors='coerce').dt.date
            except Exception:
                df_cdp["posting_date_CDP"] = pd.NaT
        else:
            df_cdp["posting_date_CDP"] = pd.NaT

        # Aplicar filtro de fecha si se especificó
        if fecha_filtro is not None:
            if posting_col:
                print(f"   Aplicando filtro de fecha desde {fecha_filtro} en columna '{posting_col}'")
                df_cdp = df_cdp[pd.to_datetime(df_cdp[posting_col], errors='coerce').dt.date >= fecha_filtro].copy()
                print(f"   Filas después del filtro: {len(df_cdp):,} (antes: {len(df_cdp):,})")
            else:
                print(f"   [AVISO] No se encontró columna 'Fecha Posting' para aplicar filtro de fecha")
    else:
        print("   CDP  no provisto, inicializando vacío...")
        df_cdp = pd.DataFrame()

    if e1_path:
        print(f"   [INFO] Archivo E1 recibido. Lógica pendiente de implementación.")

    if diagnostico:
        print("   Encabezados ARCA (crudos):", list(df_arca_raw.columns)[:30])
        print("   Encabezados CDP  (crudos):", list(df_cdp.columns)[:30] if not df_cdp.empty else [])

    # 2) Limpieza CDP (errores de Excel)
    na_like = ["#N/A", "#¡N/A!", "#N/D", "#NOMBRE?", "#VALUE!", "#VALOR!", "#REF!", "#DIV/0!", ""]
    for col in df_cdp.columns:
        df_cdp[col] = df_cdp[col].replace(na_like, pd.NA)

    # 3) Ajuste de mapping dinámico (Moneda/Monto)
    print("3/8 Ajustando mapeos (Moneda/Monto) ...")
    map_arca = dict(config["columnas"]["ARCA"])
    map_cdp  = dict(config["columnas"]["CDP"])

    # ARCA: mejor currency y amount
    mejor_moneda_arca = seleccionar_mejor_currency(df_arca_raw)
    if mejor_moneda_arca:
        for k, v in list(map_arca.items()):
            if v == "currency":
                map_arca.pop(k)
        map_arca[mejor_moneda_arca] = "currency"

    mejor_amount_arca = seleccionar_mejor_amount(df_arca_raw)
    # ARCA: Forzar siempre columna "Monto"
    for k, v in list(map_arca.items()):
        if v == "amount":
            map_arca.pop(k)
    map_arca["Monto"] = "amount"

    # CDP: forzar “Monto total” si existe; si no, seleccionar mejor monto
    # (evita que el selector elija “Monto” cuando queremos “Monto total”)
    if not df_cdp.empty and "Monto total" in df_cdp.columns:
        # eliminar cualquier mapping previo de amount y setear “Monto total”
        for k, v in list(map_cdp.items()):
            if v == "amount":
                map_cdp.pop(k)
        map_cdp["Monto total"] = "amount"
    elif not df_cdp.empty:
        mejor_amount_cdp = seleccionar_mejor_amount(df_cdp)
        if mejor_amount_cdp:
            for k, v in list(map_cdp.items()):
                if v == "amount":
                    map_cdp.pop(k)
            map_cdp[mejor_amount_cdp] = "amount"

    # CDP: Asegurar mapeo de Doc Nbr si existe
    if not df_cdp.empty:
        for c in df_cdp.columns:
            if canonicalize(c) in ["docnbr", "nrodoc"] and "nbr_doc" not in map_cdp.values():
                map_cdp[c] = "nbr_doc"

    # CDP: Asegurar mapeo de invoice_number si existe
    if not df_cdp.empty:
        for c in df_cdp.columns:
            if canonicalize(c) in ["invoicenumber", "numeroinvoice", "invoice_number", "numero_factura"] and "invoice_number" not in map_cdp.values():
                map_cdp[c] = "invoice_number"

    # 4) Renombrado robusto
    print("4/8 Renombrando columnas según YAML ...")
    req_base = ["Concatenado", "amount", "currency", "vendor", "issue_date", "duplicados"]
    df_arca_raw = rename_robusto(df_arca_raw, map_arca, "ARCA", req_base)

    req_cdp = req_base.copy()
    if "doc_type" in map_cdp.values():
        req_cdp.append("doc_type")
    # Nbr doc NO es obligatoria; si está, se arrastra igual
    
    if not df_cdp.empty:
        df_cdp = rename_robusto(df_cdp, map_cdp, "CDP", req_cdp)
    else:
        df_cdp = pd.DataFrame(columns=req_cdp)

    # Moneda por defecto CDP si viene en blanco
    default_curr = config.get("default_currency_cdp", None)
    if default_curr and "currency" in df_cdp.columns:
        df_cdp["currency"] = df_cdp["currency"].fillna(default_curr).replace("", default_curr)

    assert_cols(df_arca_raw, "ARCA", req_base)
    assert_cols(df_cdp,      "CDP",  req_cdp)

    # 5) Normalizaciones
    print("5/8 Normalizando montos/monedas/fechas ...")
    df_cdp = df_cdp[~df_cdp["Concatenado"].isna()].copy()
    if "erp" not in df_cdp.columns:
        df_cdp["erp"] = "(sin ERP)"
    if "doc_type" not in df_cdp.columns:
        df_cdp["doc_type"] = ""

    df_arca_raw = parsear_fecha_columna(df_arca_raw, "issue_date")
    df_cdp      = parsear_fecha_columna(df_cdp, "issue_date")

    df_arca_raw["amount_clean"] = convertir_monto(df_arca_raw["amount"])
    df_cdp["amount_clean"]      = convertir_monto(df_cdp["amount"])

    eq = config.get("equivalencias_moneda", {})
    df_arca_raw["currency_norm"] = df_arca_raw["currency"].apply(lambda x: normalizar_moneda(x, eq))
    df_cdp["currency_norm"]      = df_cdp["currency"].apply(lambda x: normalizar_moneda(x, eq))

    df_arca_raw["duplicados"] = normalizar_duplicados(df_arca_raw["duplicados"])
    df_cdp["duplicados"]      = normalizar_duplicados(df_cdp["duplicados"])

    # 6) Comparaciones y duplicados
    print("6/8 Calculando duplicados y comparando por Concatenado ...")
    universo_concat = set(df_cdp["Concatenado"].astype(str).unique())
    df_arca         = df_arca_raw[df_arca_raw["Concatenado"].astype(str).isin(universo_concat)].copy()

    universo_dup    = set(df_cdp["duplicados"].astype(str).unique())
    arca_dup_counts = df_arca_raw.groupby("duplicados").size().reset_index(name="ARCA_dup_count")
    arca_dup_counts = arca_dup_counts[arca_dup_counts["duplicados"].astype(str).isin(universo_dup)].copy()
    arca_dup_counts["EsDuplicado_ARCA"] = arca_dup_counts["ARCA_dup_count"] >= 2
    arca_dup_counts = arca_dup_counts.sort_values(["ARCA_dup_count", "duplicados"], ascending=[False, True])

    dup_keys = set(arca_dup_counts.loc[arca_dup_counts["ARCA_dup_count"] >= 2, "duplicados"].astype(str))
    df_duplicados_arca_detalle = df_arca_raw[df_arca_raw["duplicados"].astype(str).isin(dup_keys)].copy()

    # --- CDP: agregación (incluye nbr_doc si existe) ---
    df_cdp["CDP_line_count"] = 1
    agg_rules = {
        "amount_clean": lambda x: x.iloc[0],  # Cambiar a first para evitar duplicados en facturas con múltiples filas
        "currency_norm": "first",
        "vendor": "first",
        "issue_date": "first",
        "erp": "first",
        "CDP_line_count": "sum",
    }
    # Asegurar que la fecha de posting forme parte de la agregación (historial por posting_date)
    agg_rules["posting_date_CDP"] = "first"
    if "doc_type" in df_cdp.columns:
        agg_rules["doc_type"] = "first"
    if "tipo_cambio" in df_cdp.columns:
        agg_rules["tipo_cambio"] = "first"
    if "invoice_number" in df_cdp.columns:
        agg_rules["invoice_number"] = "first"

    if "nbr_doc" in df_cdp.columns:
        agg_rules["Concatenado"] = "first"
        df_cdp_valid = df_cdp[df_cdp["nbr_doc"].notna() & (df_cdp["nbr_doc"].astype(str).str.strip() != "")]
        df_cdp_miss = df_cdp[~(df_cdp["nbr_doc"].notna() & (df_cdp["nbr_doc"].astype(str).str.strip() != ""))]

        df_g1 = df_cdp_valid.groupby(["nbr_doc", "posting_date_CDP"], as_index=False).agg(agg_rules)

        if "nbr_doc" not in agg_rules:
            agg_rules["nbr_doc"] = "first"
        df_g2 = df_cdp_miss.groupby(["Concatenado", "posting_date_CDP"], as_index=False).agg(agg_rules)

        df_cdp_g = pd.concat([df_g1, df_g2], ignore_index=True)
    else:
        df_cdp_g = df_cdp.groupby(["Concatenado", "posting_date_CDP"], as_index=False).agg(agg_rules)

    # --- ARCA: agregación (añade tipo si existe) ---
    agg_arca = {
        "amount_clean": "sum",
        "currency_norm": "first",
        "vendor": "first",
        "issue_date": "first",
    }
    if "tipo" in df_arca.columns:
        agg_arca["tipo"] = "first"
    if "tipo_cambio" in df_arca.columns:
        agg_arca["tipo_cambio"] = "first"

    df_arca_g = df_arca.groupby("Concatenado", as_index=False).agg(agg_arca)

    # --- Claves duplicados por Concatenado
    cdp_keys_by_concat = (
        df_cdp.groupby("Concatenado")["duplicados"]
              .agg(lambda s: list(pd.unique(s)))
              .to_dict()
    )
    arca_count_by_key = dict(
        zip(arca_dup_counts["duplicados"].astype(str), arca_dup_counts["ARCA_dup_count"])
    )

    def max_arca_count_for_concat(concat_val):
        keys = cdp_keys_by_concat.get(concat_val, [])
        if not keys:
            return 0
        return max(arca_count_by_key.get(str(k), 0) for k in keys)

    # Merge
    df_merge = pd.merge(
        df_cdp_g, df_arca_g,
        on="Concatenado", suffixes=("_CDP", "_ARCA"), how="left"
    )
    df_merge["ARCA_dup_count"] = df_merge["Concatenado"].apply(max_arca_count_for_concat).astype(int)

    # >>> Mostrar filas que no existen en ARCA (visualizar faltantes)
    # df_merge = df_merge[~df_merge["amount_clean_ARCA"].isna()].copy()

    # 7) Status (incluye "Faltante en ARCA")
    print("7/8 Clasificando Status ...")
    tol = float(config.get("tolerancia_monto", 2.0))

    def _normalize_tipo_cambio(value):
        if pd.isna(value):
            return None
        if isinstance(value, str):
            text = value.strip()
            if text == "":
                return None
            text = text.replace(",", ".")
            try:
                return float(text)
            except ValueError:
                return text
        return value

    def _is_local_currency_rate(rate):
        return isinstance(rate, (int, float)) and 0.0 <= rate <= 2.0

    EXCHANGE_RATE_TOLERANCE = 0.50

    def _tipo_cambio_match(val_cdp, val_arca):
        rate_cdp = _normalize_tipo_cambio(val_cdp)
        rate_arca = _normalize_tipo_cambio(val_arca)
        if rate_cdp is None or rate_arca is None:
            return True
        if _is_local_currency_rate(rate_cdp) or _is_local_currency_rate(rate_arca):
            return True
        if isinstance(rate_cdp, (int, float)) and isinstance(rate_arca, (int, float)):
            rate_cdp = round(float(rate_cdp), 2)
            rate_arca = round(float(rate_arca), 2)
            return abs(rate_cdp - rate_arca) <= EXCHANGE_RATE_TOLERANCE
        return str(rate_cdp).strip() == str(rate_arca).strip()

    def classify(row):
        # Primero verificar si falta en ARCA
        if pd.isna(row.get("amount_clean_ARCA")):
            return "Faltante en ARCA"
        if row["ARCA_dup_count"] >= 2:
            return "Duplicado en ARCA"
        if not _tipo_cambio_match(row.get("tipo_cambio_CDP"), row.get("tipo_cambio_ARCA")):
            return "Error de Tipo Cambio"
        if row.get("currency_norm_CDP") != row.get("currency_norm_ARCA"):
            return "Error de Moneda"
        val_cdp = row.get("amount_clean_CDP", 0)
        val_arc = row.get("amount_clean_ARCA", 0)
        diff = abs(abs(val_cdp) - abs(val_arc))
        if diff > tol:
            return "Error de Monto"
        return "OK"

    df_merge["Status"] = df_merge.apply(classify, axis=1)
    df_merge = df_merge[df_merge["Status"] != "Faltante en ARCA"].copy()
    df_merge["Diferencia_Monto"] = df_merge["amount_clean_CDP"].fillna(0) - df_merge["amount_clean_ARCA"].fillna(0)

    resumen_por_estado = (
        df_merge.groupby("Status")
                .agg(Cantidad=("Concatenado", "count"))
                .reset_index()
                .sort_values("Status")
    )

    # Ajustes de columnas para el reporte
    if "vendor_ARCA" in df_merge.columns:
        df_merge = df_merge.drop(columns=["vendor_ARCA"])

    if "numero_factura" not in df_merge.columns:
        if "invoice_number_CDP" in df_merge.columns:
            df_merge["numero_factura"] = df_merge["invoice_number_CDP"]
        elif "invoice_number" in df_merge.columns:
            df_merge["numero_factura"] = df_merge["invoice_number"]
        elif "invoice_number_ARCA" in df_merge.columns:
            df_merge["numero_factura"] = df_merge["invoice_number_ARCA"]

    # Evitar duplicados de columnas de factura
    drop_invoice_cols = [c for c in ["invoice_number_CDP", "invoice_number", "invoice_number_ARCA"] if c in df_merge.columns]
    if drop_invoice_cols:
        df_merge = df_merge.drop(columns=drop_invoice_cols)

    # 8) Orden de columnas (Nbr doc 4ª)
    cols_pref = [
        "posting_date_CDP", "Status", "erp", "doc_type", "nbr_doc",
        "tipo_ARCA",
        "Concatenado",
        "amount_clean_CDP", "currency_norm_CDP", "tipo_cambio_CDP", "issue_date_CDP", "vendor_CDP",
        "CDP_line_count",
        "amount_clean_ARCA", "currency_norm_ARCA", "tipo_cambio_ARCA", "issue_date_ARCA",
        "ARCA_dup_count",
        "Diferencia_Monto",
    ]
    cols_final = [c for c in cols_pref if c in df_merge.columns] + \
                 [c for c in df_merge.columns if c not in cols_pref]
    df_merge = df_merge[cols_final]
    df_err   = df_merge[df_merge["Status"] != "OK"].copy()

    # 8) Exportar
    print("8/8 Exportando a Excel (Formatos integrados)...")
    base_path_for_report = cdp_path if cdp_path else e1_path
    if not base_path_for_report: base_path_for_report = arca_path
    output_path = Path(base_path_for_report).parent / "Reporte_Control_Resultados.xlsx"

    # Preparar columnas requeridas en el orden solicitado
    required_order = [
        "posting_date_CDP","Status","erp","doc_type","nbr_doc","numero_factura","Concatenado",
        "amount_clean_CDP","currency_norm_CDP","tipo_cambio_CDP","issue_date_CDP","vendor_CDP","CDP_line_count",
        "amount_clean_ARCA","currency_norm_ARCA","tipo_cambio_ARCA","issue_date_ARCA","ARCA_dup_count","Diferencia_Monto",
    ]

    # Normalizar nombres que puedan venir con sufijos
    if "nbr_doc_CDP" in df_merge.columns and "nbr_doc" not in df_merge.columns:
        df_merge = df_merge.rename(columns={"nbr_doc_CDP": "nbr_doc"})

    # Asegurar existencia de columnas solicitadas
    for c in required_order:
        if c not in df_merge.columns:
            df_merge[c] = pd.NA

    df_current_baseline = df_merge[required_order].copy()

    # Mantener historial horizontal: si existe reporte, cargar 'Detalle Completo' y agregar columnas con timestamp
    from datetime import datetime
    run_ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    if output_path.exists():
        try:
            df_existing = pd.read_excel(output_path, sheet_name="Detalle Completo", engine="openpyxl")
        except Exception:
            df_existing = pd.DataFrame()
    else:
        df_existing = pd.DataFrame()

    # Convertir posting_date a string para evitar problemas de key
    df_current_baseline["posting_date_CDP"] = df_current_baseline["posting_date_CDP"].astype(str)
    if not df_existing.empty:
        if "posting_date_CDP" not in df_existing.columns:
            df_existing["posting_date_CDP"] = ""
        else:
            df_existing["posting_date_CDP"] = df_existing["posting_date_CDP"].astype(str)
        if "Concatenado" not in df_existing.columns:
            df_existing["Concatenado"] = ""

    if df_existing.empty:
        # Primera corrida: escribir el baseline tal cual
        df_out = df_current_baseline
    else:
        # Versionar columnas (excepto keys)
        key_cols = []
        if "posting_date_CDP" in df_existing.columns and "posting_date_CDP" in df_current_baseline.columns:
            key_cols.append("posting_date_CDP")
        if "Concatenado" in df_existing.columns and "Concatenado" in df_current_baseline.columns:
            key_cols.append("Concatenado")
        cols_to_version = [c for c in required_order if c not in key_cols]
        df_curr_ver = df_current_baseline.copy()
        df_curr_ver = df_curr_ver.rename(columns={c: f"{c}_{run_ts}" for c in cols_to_version})

        # Hacer merge outer para conservar histórico y agregar nuevas columnas
        if key_cols:
            df_out = pd.merge(df_existing, df_curr_ver, on=key_cols, how="outer")
        else:
            df_out = pd.concat([df_existing, df_curr_ver], ignore_index=True, sort=False)

    # Escribir reporte (sin Dashboard ni Data_Analisis)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_out.to_excel(writer, sheet_name="Detalle Completo", index=False)
        df_err.to_excel(writer, sheet_name="Errores y Faltantes", index=False)
        if not df_duplicados_arca_detalle.empty:
            df_duplicados_arca_detalle.to_excel(writer, sheet_name="Duplicados ARCA", index=False)

    # === CRONÓMETRO (fin) ===
    t1 = time.perf_counter()
    elapsed = t1 - t0
    elapsed_str = str(timedelta(seconds=int(elapsed)))
    msg_exec_time = f"Tiempo total de ejecución: {elapsed:.2f} s (~ {elapsed_str})"
    print(msg_exec_time)
    logging.warning(msg_exec_time)

    # Post-proceso: aplicar estilos en hojas existentes
    wb = load_workbook(output_path)
    for hoja in ["Detalle Completo", "Errores y Faltantes", "Duplicados ARCA"]:
        if hoja in wb.sheetnames:
            ws = wb[hoja]
            aplicar_estilo_tabla(ws)
            aplicar_formato_condicional_avanzado(ws)
    wb.save(output_path)
    print(f"Reporte generado: {output_path}")
    return str(output_path), resumen_por_estado, df_err


class ControlApp:
    def __init__(self, root, config):
        self.root = root
        self.config = config
        self.root.title("Sistema de Control de Facturas (ARCA vs CDP)")
        self.root.geometry("1000x750")
        
        self.arca_path = tk.StringVar()
        self.cdp_path = tk.StringVar()
        self.e1_path = tk.StringVar()
        self.fecha_posting = tk.StringVar()
        self.fecha_posting.set("Todas las fechas")
        
        self.create_widgets()
        
    def create_widgets(self):
        style = ttk.Style()
        style.configure("TButton", font=("Segoe UI", 10))
        style.configure("TLabel", font=("Segoe UI", 10))
        
        file_frame = ttk.LabelFrame(self.root, text="Selección de Archivos", padding=15)
        file_frame.pack(fill=tk.X, padx=15, pady=10)
        
        ttk.Button(file_frame, text="📁 Seleccionar Archivo ARCA", command=self.select_arca, width=30).grid(row=0, column=0, padx=5, pady=5)
        ttk.Label(file_frame, textvariable=self.arca_path, width=70, relief="sunken", padding=2).grid(row=0, column=1, sticky=tk.W, padx=10)
        
        ttk.Button(file_frame, text="📁 Seleccionar Archivo CDP", command=self.select_cdp, width=30).grid(row=1, column=0, padx=5, pady=5)
        ttk.Label(file_frame, textvariable=self.cdp_path, width=70, relief="sunken", padding=2).grid(row=1, column=1, sticky=tk.W, padx=10)
        
        ttk.Button(file_frame, text="📁 Seleccionar Archivo E1", command=self.select_e1, width=30).grid(row=2, column=0, padx=5, pady=5)
        ttk.Label(file_frame, textvariable=self.e1_path, width=70, relief="sunken", padding=2).grid(row=2, column=1, sticky=tk.W, padx=10)
        
        # Filtro de fecha
        filter_frame = ttk.LabelFrame(self.root, text="Filtro de Fecha", padding=15)
        filter_frame.pack(fill=tk.X, padx=15, pady=10)
        
        ttk.Label(filter_frame, text="Fecha Posting (desde):", width=20).grid(row=0, column=0, padx=5, pady=5)
        self.fecha_combo = ttk.Combobox(filter_frame, textvariable=self.fecha_posting, state="readonly", width=20)
        self.fecha_combo.grid(row=0, column=1, padx=5, pady=5)
        self.fecha_combo['values'] = ["Todas las fechas"]
        
        btn_frame = ttk.Frame(self.root)
        btn_frame.pack(pady=5)
        
        self.btn_run = ttk.Button(btn_frame, text="▶ Ejecutar Control", command=self.run_process_thread, width=25)
        self.btn_run.pack(side=tk.LEFT, padx=5)
        
        self.btn_open = ttk.Button(btn_frame, text="📄 Abrir reporte", command=self.open_report, width=25, state=tk.DISABLED)
        self.btn_open.pack(side=tk.LEFT, padx=5)
        
        self.status_var = tk.StringVar()
        self.status_var.set("Esperando...")
        
        self.progress = ttk.Progressbar(self.root, mode='determinate', length=500)
        self.progress.pack(pady=5)
        
        ttk.Label(self.root, textvariable=self.status_var, font=("Segoe UI", 9, "italic")).pack(pady=(0, 10))
        
        content_frame = ttk.Frame(self.root)
        content_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=10)
        
        table_frame = ttk.LabelFrame(content_frame, text="Registros con Errores", padding=5)
        table_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        
        self.tree_scroll_y = ttk.Scrollbar(table_frame)
        self.tree_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.tree_scroll_x = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL)
        self.tree_scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        
        self.tree = ttk.Treeview(table_frame, yscrollcommand=self.tree_scroll_y.set, xscrollcommand=self.tree_scroll_x.set, selectmode="extended")
        self.tree.pack(fill=tk.BOTH, expand=True)
        
        self.tree_scroll_y.config(command=self.tree.yview)
        self.tree_scroll_x.config(command=self.tree.xview)
        
        sys.stdout = ProgressRedirector(self.status_var, self.progress, self.root)
        
        self.chart_frame = ttk.LabelFrame(content_frame, text="Resumen Gráfico", width=400)
        self.chart_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=False, padx=(10, 0))
        self.chart_frame.pack_propagate(False)
        
    def select_arca(self):
        path = filedialog.askopenfilename(title="Seleccionar Archivo ARCA", filetypes=[("Excel", "*.xlsx *.xls")])
        if path:
            self.arca_path.set(path)
            
    def select_cdp(self):
        path = filedialog.askopenfilename(title="Seleccionar Archivo CDP", filetypes=[("Excel", "*.xlsx *.xls")])
        if path:
            self.cdp_path.set(path)
            self.load_fechas_posting(path)
            
    def select_e1(self):
        path = filedialog.askopenfilename(title="Seleccionar Archivo E1", filetypes=[("Excel", "*.xlsx *.xls")])
        if path:
            self.e1_path.set(path)
            
    def load_fechas_posting(self, cdp_path):
        def _is_posting_column(col_name):
            col_lower = str(col_name).lower()
            return "posting" in col_lower and ("fecha" in col_lower or "date" in col_lower or "posting" in col_lower)

        try:
            xls = pd.ExcelFile(cdp_path, engine="openpyxl")
            sheet_names = xls.sheet_names
            candidate_sheets = [sheet for sheet in ["CDP"] + sheet_names if sheet in sheet_names]

            col_fecha_posting = None
            sheet_con_found = None
            for sheet in candidate_sheets:
                try:
                    df_temp = pd.read_excel(cdp_path, sheet_name=sheet, nrows=0, engine="openpyxl")
                except Exception:
                    continue
                for col in df_temp.columns:
                    if _is_posting_column(col):
                        col_fecha_posting = col
                        sheet_con_found = sheet
                        break
                if col_fecha_posting:
                    break

            if col_fecha_posting:
                df_fechas = pd.read_excel(cdp_path, sheet_name=sheet_con_found, usecols=[col_fecha_posting], engine="openpyxl")
                fechas = pd.to_datetime(df_fechas[col_fecha_posting], errors='coerce').dropna().dt.date.unique()
                fechas_ordenadas = sorted(fechas)

                opciones = ["Todas las fechas"] + [fecha.strftime("%Y-%m-%d") for fecha in fechas_ordenadas]
                self.fecha_combo['values'] = opciones
                self.fecha_posting.set("Todas las fechas")
                print(f"Fechas de posting cargadas: {len(fechas_ordenadas)} fechas únicas encontradas")
            else:
                print(f"Advertencia: No se encontró columna de posting date en el archivo CDP")
                self.fecha_combo['values'] = ["Todas las fechas"]
                self.fecha_posting.set("Todas las fechas")

        except Exception as e:
            print(f"Error al cargar fechas de posting: {e}")
            self.fecha_combo['values'] = ["Todas las fechas"]
            self.fecha_posting.set("Todas las fechas")
            
    def run_process_thread(self):
        if not self.arca_path.get() or (not self.cdp_path.get() and not self.e1_path.get()):
            messagebox.showwarning("Advertencia", "Debe seleccionar el archivo ARCA y al menos uno de los ERPs (CDP o E1).")
            return
            
        self.btn_run.config(state=tk.DISABLED)
        self.btn_open.config(state=tk.DISABLED)
        self.progress["value"] = 0
        self.status_var.set("Iniciando proceso en 2do plano...")
        self.tree.delete(*self.tree.get_children())
        
        for widget in self.chart_frame.winfo_children():
            widget.destroy()
            
        thread = threading.Thread(target=self.run_process)
        thread.daemon = True
        thread.start()
        
    def run_process(self):
        try:
            # Obtener la fecha seleccionada
            fecha_filtro = None
            if self.fecha_posting.get() != "Todas las fechas":
                from datetime import datetime
                fecha_filtro = datetime.strptime(self.fecha_posting.get(), "%Y-%m-%d").date()
            
            out_path, df_resumen, df_err = process_data(self.config, self.arca_path.get(), self.cdp_path.get(), self.e1_path.get(), fecha_filtro)
            self.root.after(0, self.process_success, out_path, df_resumen, df_err)
        except Exception as e:
            logging.error("Error crítico en procesamiento", exc_info=True)
            self.root.after(0, self.process_error, str(e))
            
    def process_success(self, out_path, df_resumen, df_err):
        self.progress["value"] = 100
        self.status_var.set(f"Proceso finalizado. Reporte en: {Path(out_path).name}")
        self.btn_run.config(state=tk.NORMAL)
        
        self.last_out_path = out_path
        self.btn_open.config(state=tk.NORMAL)
        
        self.draw_chart(df_resumen)
        self.populate_treeview(df_err)
        
        try:
            import os
            os.startfile(Path(out_path).parent)
        except Exception:
            pass
        messagebox.showinfo("Éxito", "El proceso finalizó correctamente.")

    def open_report(self):
        if hasattr(self, 'last_out_path') and self.last_out_path:
            try:
                import os
                os.startfile(self.last_out_path)
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo abrir el reporte:\n{e}")

    def populate_treeview(self, df_err):
        self.tree.delete(*self.tree.get_children())
        if df_err.empty:
            return
            
        columns = list(df_err.columns)
        self.tree["columns"] = columns
        self.tree["show"] = "headings"
        
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=100, minwidth=50)
            
        for _, row in df_err.iterrows():
            vals = [str(x) if not pd.isna(x) else "" for x in row.tolist()]
            self.tree.insert("", tk.END, values=vals)

    def process_error(self, err_msg):
        print(f"\n[ERROR] Ocurrió un error:\n{err_msg}")
        self.btn_run.config(state=tk.NORMAL)
        messagebox.showerror("Error", f"Ocurrió un error en la ejecución:\n{err_msg}")
        
    def draw_chart(self, df_resumen):
        if df_resumen.empty:
            ttk.Label(self.chart_frame, text="Sin datos para graficar").pack(pady=20)
            return
            
        colors_cfg = self.config.get("dashboard_colors", {})
        default_colors = {"OK": "00B050", "Error de Monto": "FF99CC", "Error de Moneda": "FFFF00", "Duplicado en ARCA": "D6EAF8"}
        
        labels = df_resumen["Status"].tolist()
        sizes = df_resumen["Cantidad"].tolist()
        colors = ["#" + colors_cfg.get(st, default_colors.get(st, "CCCCCC")).replace("#", "") for st in labels]
        
        fig = Figure(figsize=(4, 4), dpi=100)
        ax = fig.add_subplot(111)
        ax.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', startangle=140)
        ax.axis('equal')
        fig.tight_layout()
        
        canvas = FigureCanvasTkAgg(fig, master=self.chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

def main():
    try:
        config = cargar_config()
        setup_logging(config.get("logging", {}).get("level", "WARNING"))
        
        root = tk.Tk()
        app = ControlApp(root, config)
        root.mainloop()
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        input("Presiona Enter para salir...")

if __name__ == "__main__":
    main()