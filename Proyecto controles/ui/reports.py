"""Módulo de generación de reportes Excel."""
import pandas as pd
import logging
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.styles import Font, PatternFill, Alignment
from ui.styles import (
    COLOR_HEADER_BG, COLOR_OK_FILL, COLOR_ERROR_FILL, 
    ESTILO_BORDE_TABLA, apply_header_style, get_status_color
)


logger = logging.getLogger(__name__)


def generate_dashboard(wb, df_resumen, df_detalle, config, exec_msg=None):
    """Genera el dashboard visual en una hoja de Excel."""
    
    if "Dashboard" in wb.sheetnames:
        ws = wb["Dashboard"]
    else:
        ws = wb.create_sheet("Dashboard", 0)
    
    ws.sheet_view.showGridLines = False
    
    # Título
    ws["B2"] = "DASHBOARD DE CONTROL DE FACTURAS"
    ws["B2"].font = Font(size=20, bold=True, color=COLOR_HEADER_BG)
    ws.merge_cells("B2:J2")
    
    # KPIs
    total_regs = len(df_detalle)
    total_ok = len(df_detalle[df_detalle["Status"] == "OK"])
    total_err = total_regs - total_ok
    pct_ok = (total_ok / total_regs) if total_regs else 0
    
    kpi_start_row = 4
    kpis = [
        ("Total Registros", total_regs, "E8DAEF"),
        ("Correctos (OK)", total_ok, COLOR_OK_FILL),
        ("Con Errores/Alertas", total_err, COLOR_ERROR_FILL),
        ("% Efectividad", f"{pct_ok:.1%}", "D6EAF8"),
    ]
    
    col = 2
    for title, val, color in kpis:
        cell_title = ws.cell(row=kpi_start_row, column=col)
        cell_val = ws.cell(row=kpi_start_row + 1, column=col)
        
        cell_title.value = title
        cell_title.font = Font(bold=True, size=12, color="555555")
        cell_title.alignment = Alignment(horizontal="center")
        
        cell_val.value = val
        cell_val.font = Font(bold=True, size=16)
        cell_val.alignment = Alignment(horizontal="center")
        cell_val.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        for r in range(kpi_start_row, kpi_start_row + 2):
            ws.cell(row=r, column=col).border = ESTILO_BORDE_TABLA
        
        col += 2
    
    # Tiempo de ejecución
    if exec_msg:
        ws["B6"] = "Tiempo de ejecución"
        ws["C6"] = exec_msg
        ws["B6"].font = Font(bold=True, color="555555")
        ws["C6"].font = Font(bold=False, color="555555")
    
    # Tabla de resumen
    ws["B8"] = "Resumen por Estado"
    ws["B8"].font = Font(bold=True, size=14, color=COLOR_HEADER_BG)
    
    start_row_data = 9
    ws.cell(row=start_row_data, column=2, value="Estado")
    ws.cell(row=start_row_data, column=3, value="Cantidad")
    
    for idx, row in df_resumen.iterrows():
        r = start_row_data + 1 + idx
        ws.cell(row=r, column=2, value=row["Status"])
        ws.cell(row=r, column=3, value=row["Cantidad"])
    
    apply_header_style(ws, start_row_data, start_row_data + len(df_resumen), 2, 3)
    
    # Gráfico circular
    pie = PieChart()
    pie.title = "Distribución de Estados"
    pie.width = 15
    pie.height = 10
    
    labels = Reference(ws, min_col=2, min_row=start_row_data + 1, max_row=start_row_data + len(df_resumen))
    data = Reference(ws, min_col=3, min_row=start_row_data, max_row=start_row_data + len(df_resumen))
    
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    
    colors_cfg = config.get("dashboard_colors", {})
    
    for i, status in enumerate(df_resumen["Status"]):
        color_hex = colors_cfg.get(status, {}).replace("#", "") if isinstance(colors_cfg.get(status), str) else "CCCCCC"
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color_hex
        pie.series[0].dPt.append(pt)
    
    ws.add_chart(pie, "E8")
    
    # Gráfico de barras
    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "Cantidad por Estado"
    bar.y_axis.title = 'Registros'
    bar.x_axis.title = 'Estado'
    bar.width = 15
    bar.height = 10
    
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(labels)
    
    for i, status in enumerate(df_resumen["Status"]):
        color_hex = colors_cfg.get(status, {}).replace("#", "") if isinstance(colors_cfg.get(status), str) else "CCCCCC"
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color_hex
        bar.series[0].dPt.append(pt)
    
    ws.add_chart(bar, "L8")
